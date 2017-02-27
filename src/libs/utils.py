#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''
@author: King
@deprecated: 2012-2-29
'''

from openpyxl.reader.excel import load_workbook
import time
import types
import os
import sys
import locale
import platform
import inspect
import glob
reload(sys)
sys.setdefaultencoding('utf-8') #IGNORE:E1101
locale.setlocale(locale.LC_ALL, "")


dict_var = {}

def replace_quote(data):
    return format_float(data).replace('"', '')

def format_float(data):
    '''
    格式化浮点数，主要用于去掉最右边的0和小数点
    @param data: number
    @return: string
    '''
    data = str(data)
    if data.endswith(".0"):
        return data.rstrip("0").rstrip(".")
    return data

def run_time(func):
    '''
    计算运行时间
    @param func: 函数名，装饰器模式，无需传值
    '''
    current_time = lambda tips, fun_name, cost_time=0:"[%s]-[%s]-%s-run:%.2f's" % (time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),
                                                                       tips, fun_name, cost_time)
    def __wrapper(*args, **kwargs):
        '''
        装饰模式核心
        '''
        start_time = time.clock()
        print(current_time("Start", func.__name__))
        func(*args, **kwargs)
        print(current_time("End", func.__name__, time.clock() - start_time))
    return __wrapper

def win2unix(fname):
    '''
    convert format file from windows to unix
    @param fname: the name of file
    @return: 0: Successfully!
            -1: This is Not a binary file.
             1: Don't convert this file.
    '''
    data = open(fname, "rb").read()
    if '\0' in data:
        return -1
    newdata = data.replace("\r\n", "\n")
    if newdata != data:
        open(fname, "wb").write(newdata)
        return 0
    return 1

def load_excel(excel_name):
    '''
    加载excel
    @param excel_name: Excel名称，不需要后缀，目前仅支持office 2007
    @return: work_book
    '''
    filename = ur"ext_docs/{0}.xlsx".format(excel_name) if platform.system() == "Windows" else "ext_docs/{0}.xlsx".format(excel_name)
    return load_workbook(filename, use_iterators = True)

def get_value(value, default):
    '''
    检查value是不是None,如果是的话，使用default代替
    '''
    if value == None:
        return default
    else:
        return value

def prev(content, row, col):
    '''
    从当前行开始, 向上搜索到第一个不是None的数据
    为了处理合并单元格的操作
    '''
    while row > 1:
        if content[row][col] == None:
            row = row - 1
            continue
        else:
            break
    return content[row][col]

def get_sheet_list(excel_name):
    '''
    获取一个excel的sheet名列表
    '''
    wb = load_excel(excel_name)
    return wb.get_sheet_names()

class load_sheel(object):
    '''
    加载shell数据
    @param work_book: 工作薄对象
    @param sheel_name: sheel名称
    '''
    def __init__(self, work_book, sheel_name, vars_list=[]):
        self.work_book = work_book
        self.sheel_name = sheel_name
        self.vars_list = vars_list
        self.is_exists_var = lambda var: var in globals().keys()

    def __call__(self, func):
        def __wrapper():
            ws = self.work_book.get_sheet_by_name(name=self.sheel_name)
            result = []
            is_set_title = False
            ## 获取所有的表格数据
            all_content = []
            ## 加一行空行是为了在 prev() 中不用 -1
            all_content.append([])
            for row in ws.iter_rows():
                content = []
                for cell in row:
                    content.append(cell.internal_value)
                all_content.append(content)

            for row in ws.iter_rows():
                local_row = 0   ## 当前要处理的行, 从1开始
                content = []
                for cell in row:
                    local_row = int(cell.row)
                    if cell.row == 1: continue
                    if self.vars_list and not is_set_title and cell.row == 2:
                        is_set_title = True
                    content.append(cell.internal_value)

                if content:
                    if self.vars_list and is_set_title:
                        self.set_vars(content)
                        is_set_title = False
                        continue
                    _fargs, _fvarargs, _fkeywords, _fdefaults = inspect.getargspec(func)
                    if len(_fargs) == 3:
                        result.extend(func(content, all_content, local_row))
                    elif len(_fargs) == 1:
                        result.extend(func(content))
                    else:
                    	exit("load_sheel illegal callback function!")
            self.clear_vars()
            return result
        return __wrapper

    def set_vars(self, content):
        for var in self.vars_list:
            if not self.is_exists_var(var):
                for index, excel_var in enumerate(list(content)):
                    if excel_var and excel_var == var:
                        dict_var[index] = var
                        globals()[var] = None

    def clear_vars(self):
        if self.vars_list:
            dict_var.clear()
            for var in globals().keys():
                if var in self.vars_list:
                    del globals()[var]

def module_header(module_doc, module_name, author_name="King", module_xls="TODO 请增加对应的配置表", py_name="TODO 请增加对应的源py文件名称"):
    '''
    module_header
    @param module_doc: 模块简要说明
    @param module_name: 模块名称，用于生成文件使用
    @param author_name: 作者, 默认为King
    @return: list 模块头说明
    '''
    now_time = time.strftime("%Y-%m-%d %H:%M:%S")
    content = []
    content.append("%% -------------------------------------------------------------------")
    content.append("%% @doc: {0}".format(module_doc))
    content.append("%% @author: {0}".format(author_name))
    #content.append("%% @date: {0}".format(now_time))
    content.append("%% @note: 自动生成的，请不要手动修改")
    content.append("%% 对应配置表: {0}".format(module_xls))
    content.append("%% 对应python源文件: {0}".format(py_name))
    content.append("%% @end")
    content.append("%% -------------------------------------------------------------------")
    if module_name != "":
		content.append("-module({0}).".format(module_name))
    return content

def module_hrl_header(module_doc, author_name="King", module_xls="TODO 请增加对应的配置表", py_name="TODO 请增加对应的源py文件名称"):
    '''
    module_hrl_header
    @param module_doc: 模块简要说明
    @param module_name: 模块名称，用于生成文件使用
    @param author_name: 作者, 默认为King
    @return: list 模块头说明
    '''
    now_time = time.strftime("%Y-%m-%d %H:%M:%S")
    content = []
    content.append("%% -------------------------------------------------------------------")
    content.append("%% @doc: {0}".format(module_doc))
    content.append("%% @author: {0}".format(author_name))
    #content.append("%% @date: {0}".format(now_time))
    content.append("%% @note: 自动生成的，请不要手动修改")
    content.append("%% 对应配置表: {0}".format(module_xls))
    content.append("%% 对应python源文件: {0}".format(py_name))
    content.append("%% @end")
    content.append("%% -------------------------------------------------------------------")
    return content
    
def module_lua_header(module_doc, module_name, author_name="King", module_xls="TODO 请增加对应的配置表", py_name="TODO 请增加对应的源py文件名称"):
	'''
    module_hrl_header
    @param module_doc: 模块简要说明
    @param module_name: 模块名称，用于生成文件使用
    @param author_name: 作者, 默认为King
    @return: list 模块头说明
    '''
	now_time = time.strftime("%Y-%m-%d %H:%M:%S")
	content = []
	content.append("--[[")
	content.append("@doc: {0}".format(module_doc))
	content.append("@author: {0}".format(author_name))
	content.append("@note: 自动生成的，请不要手动修改")
	content.append("对应配置表: {0}".format(module_xls))
	content.append("对应python源文件: {0}".format(py_name))
	content.append("@end")
	content.append("--]]")
	content.append("""Config = Config or {{}}
Config.{0} = {{""".format(module_name))
	return content

#@run_time
def gen_erl(module_name, data):
    '''
    生成erlang文件
    @param module_name: 模板名称
    @param data: 数据，以list方式传入
    '''
    fname = "ext_erl/{0}.erl".format(module_name)
    gen_file(fname, data)
    
#@run_time
def gen_hrl(module_name, data):
    '''
    生成erlang头文件
    @param module_name: 模板名称
    @param data: 数据，以list方式传入
    '''
    fname = "ext_hrl/{0}.hrl".format(module_name)
    gen_file(fname, data)    

#@run_time
def gen_xml(xml_name, data):
    '''
    生成xml文件
    @param xml_name: xml文件名
    @param data: 数据，以list方式传入
    '''
    fname = "ext_xml/{0}.xml".format(xml_name)
    xml_data = ["""<?xml version="1.0" encoding="utf-8"?>
<resultset xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">"""]
    xml_data.extend(data)
    xml_data.append("</resultset>")
    gen_file(fname, xml_data)

def gen_lua(lua_name, data):
    '''
    生成lua文件
    @param xml_name: xml文件名
    @param data: 数据，以list方式传入
    '''
    fname = "ext_lua/config_{0}.lua".format(lua_name)
    data.append("}")
    gen_file(fname, data)
    parse_lua_files()
    
def gen_file(fname, data):
    '''
    生成文件
    @param module_name: 模板名称
    @param data: 数据，以list方式传入
    '''    
    assert type(data) is types.ListType
    save_path = os.path.dirname(os.path.realpath(fname))
    if not os.path.exists(save_path):
        os.makedirs(save_path)
        
    open(fname, "w").write("\n".join(data))
    win2unix(fname)
    if fname[0:3] == "erl" or fname[0:3] == "hrl" or fname[0:3] == "lua":
        print("OK! <a href=\"DataConvert/view.php?file={0}\" target=\"_blank\">{0}</a>".format(fname))
    else:
        print("OK! <a href=\"DataConvert/{0}\" target=\"_blank\">{0}</a>".format(fname))    
    
## 生成lua配置，需要重新更新config.lua文件内容把对应的配置加进去
def parse_lua_files():
	file_list = []
	for file in glob.glob(r'ext_lua/config_*.lua'):
		filename = os.path.basename(file)[:-4] #去掉路径和后缀,只获取文件名
		#print filename
		file_list.append("require(\"config.{0}\")".format(filename))
	gen_file("ext_lua/config.lua", file_list)
	